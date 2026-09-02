"""
Kardin '3. Recoveries Back-up' bucket parser (v1).

Four inputs:
  - Recovery Calc Est   (PDF: per suite/reimb-type CAM+Tax calc, plus a
                          portfolio-level CAM/Tax/Total summary on page 2)
  - Recovery Monthly    (PDF: two schedules - CAM and Tax - each per suite
                          monthly $ x12 + total, mirrors bucket 1's format)
  - Gross Up Schedule    (PDF: one page per GL line's gross-up calculation,
                          grouped under a named GU group e.g. "...95% GU")
  - Fixed Factor Calcs   (xlsx: the underlying gross-up workpaper - three
                          stacked tables on the 'Modex' sheet, one per GU %)

Checks:
  1. Gross Up Schedule PDF completeness: its own header states "Page N of M"
     - if the file has fewer than M physical pages, it's missing content.
  2. Recovery Calc Est's CAM/Tax totals (page 2 summary) vs Recovery
     Monthly's CAM/Tax "Total Reimbursement" - should match exactly.
  3. Fixed Factor Calcs (the backup workpaper) vs Gross Up Schedule PDF
     (what's actually in Kardin) - per GL line, do Recoverable Expenses and
     Gross-up Amount tie out? A mismatch usually means the workpaper wasn't
     updated after Kardin's numbers changed.
"""
import re

from kardin_parser import MONEY_RE, DECIMAL_RE, is_boilerplate, to_money

DATE_RE = re.compile(r'^\d{1,2}/\d{1,2}/\d{4}$')
SUITE_RE = re.compile(r'^[Ww]est\d+-\d+$')
PAGE_HEADER_RE = re.compile(r'Page:\s*(\d+)\s+of\s+(\d+)')


def all_lines_and_page_count(pdf_file):
    import pdfplumber
    lines = []
    with pdfplumber.open(pdf_file) as pdf:
        n_pages = len(pdf.pages)
        for page in pdf.pages:
            text = page.extract_text() or ''
            lines.extend(text.split('\n'))
    return lines, n_pages


# ------------------------------------------------------------ Recovery Monthly

def parse_recovery_monthly(pdf_file):
    """Returns list of {reimb_type, suite, label, months[12], total, is_total}."""
    lines, _ = all_lines_and_page_count(pdf_file)
    rows = []
    reimb_type = None
    for raw in lines:
        line = raw.rstrip()
        if is_boilerplate(line):
            continue
        if 'CAM Recovery Schedule' in line:
            reimb_type = 'CAM'
            continue
        if 'Tax Recovery Schedule' in line:
            reimb_type = 'Tax'
            continue
        toks = line.split()
        if not toks:
            continue
        if toks[0] == 'Total' and 'Reimbursement:' in line:
            values = toks[-14:]
            if len(values) == 14 and all(MONEY_RE.match(t) for t in values[:13]) and DECIMAL_RE.match(values[13]):
                months = [to_money(t) for t in values[:12]]
                rows.append({'reimb_type': reimb_type, 'suite': None, 'label': 'TOTAL',
                             'months': months, 'total': to_money(values[12]), 'is_total': True})
            continue
        if SUITE_RE.match(toks[0]):
            date_positions = [i for i, t in enumerate(toks) if DATE_RE.match(t)]
            consecutive_dates = next(
                (i for i in date_positions if i + 1 < len(toks) and DATE_RE.match(toks[i + 1])), None)
            if consecutive_dates is None:
                continue
            i = consecutive_dates
            before = toks[1:i]
            rsf_tok = next((t for t in reversed(before) if MONEY_RE.match(t)), None)
            if rsf_tok is None:
                continue
            label = ' '.join(t for t in before if t != rsf_tok).strip()
            values = toks[i + 2:]
            if len(values) != 14 or not all(MONEY_RE.match(t) for t in values[:13]) or not DECIMAL_RE.match(values[13]):
                continue
            months = [to_money(t) for t in values[:12]]
            rows.append({'reimb_type': reimb_type, 'suite': toks[0], 'label': label,
                         'months': months, 'total': to_money(values[12]), 'is_total': False})
    return rows


# --------------------------------------------------------- Recovery Calc Est

def parse_recovery_calc_est(pdf_file):
    """
    Returns per suite/reimb-type rows: {suite, reimb_type, total_reimb} plus
    the portfolio-level summary rows: {reimb_type: 'CAM'|'Tax'|'TOTAL', total_reimb}.
    Only extracts what's needed for the cross-file tie-out checks, not the
    full calc detail (base year, %share, etc).
    """
    lines, _ = all_lines_and_page_count(pdf_file)
    detail_rows = []
    summary_rows = []
    current_suite = None
    for raw in lines:
        line = raw.rstrip()
        if is_boilerplate(line):
            continue
        toks = line.split()
        if not toks:
            continue
        if SUITE_RE.match(toks[0]):
            current_suite = toks[0]
        # per-suite CAM/Tax detail row: [.... prefix ....] MONEY MONEY DECIMAL MONEY MONEY MONEY DECIMAL
        if len(toks) >= 7:
            tail = toks[-7:]
            if (MONEY_RE.match(tail[0]) and MONEY_RE.match(tail[1]) and DECIMAL_RE.match(tail[2]) and
                    MONEY_RE.match(tail[3]) and MONEY_RE.match(tail[4]) and MONEY_RE.match(tail[5]) and
                    DECIMAL_RE.match(tail[6])):
                reimb_type = 'CAM' if 'CAM' in toks[:-7] else ('Tax' if 'Tax' in toks[:-7] else None)
                if reimb_type and current_suite:
                    detail_rows.append({'suite': current_suite, 'reimb_type': reimb_type,
                                        'total_reimb': to_money(tail[3])})
                continue
        # portfolio summary row: "CAM 1,509,570 0 1,509,570" / "Tax ..." / bare "1,706,347 0 1,706,347"
        if toks[0] in ('CAM', 'Tax') and len(toks) == 4 and all(MONEY_RE.match(t) for t in toks[1:]):
            summary_rows.append({'reimb_type': toks[0], 'total_reimb': to_money(toks[1]),
                                 'free_reimb': to_money(toks[2]), 'net_reimb': to_money(toks[3])})
        elif len(toks) == 3 and all(MONEY_RE.match(t) for t in toks):
            summary_rows.append({'reimb_type': 'TOTAL', 'total_reimb': to_money(toks[0]),
                                 'free_reimb': to_money(toks[1]), 'net_reimb': to_money(toks[2])})
    return detail_rows, summary_rows


# -------------------------------------------------------------- Gross Up Schedule

GU_BLOCK_HEADER_RE = re.compile(r'^(?P<group>.+?)\s+Gross Up Method:\s*(?P<method>\S+)\s+Occupied RSF:\s*[\d,]+$')
GL_LABEL_RE = re.compile(r'^(?P<gl>\d{6})-(?P<label>.+)$')
FIELD_RE = re.compile(r'^(?P<key>[A-Za-z /.%\-]+?):\s*(?P<val>-?[\d,]+(?:\.\d+)?%?)\s*$')
GU_TOTAL_RE = re.compile(r'^Total Gross Up Amount for:\s*(?P<group>.+?)\s*:\s*(?P<amt>[\d,.\-]+)$')


def parse_gross_up_schedule(pdf_file):
    """
    Returns (blocks, group_totals, page_info).
    blocks: [{group, gl, label, recoverable_expenses, fixed_amount_not_subject,
              amount_subject_to_gu, gu_pct, gu_amount}]
    page_info: {'stated_total_pages': int or None, 'actual_pages': int}
    """
    lines, actual_pages = all_lines_and_page_count(pdf_file)
    blocks = []
    group_totals = []
    stated_total_pages = None
    current = None
    for raw in lines:
        line = raw.rstrip()
        m_page = PAGE_HEADER_RE.search(line)
        if m_page:
            stated_total_pages = max(stated_total_pages or 0, int(m_page.group(2)))
        if is_boilerplate(line):
            continue
        m_header = GU_BLOCK_HEADER_RE.match(line)
        if m_header:
            if current:
                blocks.append(current)
            current = {'group': m_header.group('group').strip(), 'gl': None, 'label': None,
                       'recoverable_expenses': None, 'fixed_amount_not_subject': None,
                       'amount_subject_to_gu': None, 'gu_pct': None, 'gu_amount': None}
            continue
        m_total = GU_TOTAL_RE.match(line.strip())
        if m_total:
            group_totals.append({'group': m_total.group('group').strip(),
                                 'amount': float(m_total.group('amt').replace(',', ''))})
            continue
        if current is None:
            continue
        m_gl = GL_LABEL_RE.match(line.strip())
        if m_gl and current['gl'] is None:
            current['gl'] = m_gl.group('gl')
            current['label'] = m_gl.group('label').strip()
            continue
        for key, val in re.findall(r'([A-Za-z /.%\-]+?):\s*(-?[\d,]+(?:\.\d+)?%?)', line):
            key = key.strip()
            val_clean = val.replace(',', '').rstrip('%')
            try:
                num = float(val_clean)
            except ValueError:
                continue
            if key == 'Recoverable Expenses':
                current['recoverable_expenses'] = num
            elif key == 'Fixed Amount Not Subject to GU':
                current['fixed_amount_not_subject'] = num
            elif key == 'Amount Subject to Gross-Up':
                current['amount_subject_to_gu'] = num
            elif key == 'Gross Up %':
                current['gu_pct'] = num / 100.0
            elif key == 'Gross Up Amount':
                current['gu_amount'] = num
    if current:
        blocks.append(current)
    page_info = {'stated_total_pages': stated_total_pages, 'actual_pages': actual_pages}
    return blocks, group_totals, page_info


# ---------------------------------------------------------- Fixed Factor Calcs

def parse_fixed_factor_workbook(xlsx_file, sheet_name='Modex'):
    """
    Returns list of {gu_group_pct, gl_description, total_recoverable_expenses,
    fix_factor_pct, fix_factor_dollar, net_budget, gross_up_amount,
    grossed_up_adjustment, total_annual_cost, comments}.
    The sheet stacks 3 tables (one per GU %, e.g. 1.0/0.95/0.9), each preceded
    by a lone numeric row holding that table's GU % and a repeated header row.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_file, data_only=True)
    ws = wb[sheet_name]
    rows = []
    current_pct = None
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(3, 16)]  # columns C..O
        if all(v is None for v in vals):
            continue
        c0 = vals[0]
        if isinstance(c0, (int, float)) and all(v is None for v in vals[1:]):
            current_pct = float(c0)
            continue
        if c0 in ('GL Description', 'Total', '[Add Row]', None):
            continue
        gl_description = c0
        total_recoverable = vals[4]
        fix_factor_pct = vals[5]
        fix_factor_dollar = vals[6]
        net_budget = vals[7]
        gross_up_amount = vals[8]
        grossed_up_adj = vals[9]
        total_annual_cost = vals[11]
        comments = vals[12]
        if total_recoverable is None:
            continue
        rows.append({
            'gu_group_pct': current_pct, 'gl_description': gl_description,
            'total_recoverable_expenses': total_recoverable, 'fix_factor_pct': fix_factor_pct,
            'fix_factor_dollar': fix_factor_dollar, 'net_budget': net_budget,
            'gross_up_amount': gross_up_amount, 'grossed_up_adjustment': grossed_up_adj,
            'total_annual_cost': total_annual_cost, 'comments': comments,
        })
    return rows


# --------------------------------------------------------------------- Checks

def check_gross_up_schedule_completeness(page_info):
    stated, actual = page_info['stated_total_pages'], page_info['actual_pages']
    if stated is None or stated <= actual:
        return []
    return [{
        'Report Section': '7. Recoveries & Fixed Factor',
        'GL Acct': '', 'Line Item': 'Gross Up Schedule', 'Budget Year': 'Next Year Budget',
        'Priority': 'Must Fix',
        'Comment': (
            f"The Gross Up Schedule PDF's own header states 'Page {actual} of {stated}', but the file "
            f"only contains {actual} physical page(s). Pages {actual + 1}-{stated} are missing - likely "
            "a second gross-up group (e.g. a different building or GU%) or additional GL lines. "
            "Request the complete export from the PM."
        ),
        'Status': 'Open', 'Source Check': 'Incomplete Gross Up Schedule export',
    }]


def check_calc_est_vs_monthly_totals(summary_rows, monthly_rows):
    findings = []
    monthly_totals = {r['reimb_type']: r['total'] for r in monthly_rows if r['is_total']}
    for s in summary_rows:
        if s['reimb_type'] not in ('CAM', 'Tax'):
            continue
        m_total = monthly_totals.get(s['reimb_type'])
        if m_total is None:
            continue
        if s['total_reimb'] != m_total:
            findings.append({
                'Report Section': '3. Variance Comments',
                'GL Acct': '', 'Line Item': f"{s['reimb_type']} Recovery", 'Budget Year': 'Next Year Budget',
                'Priority': 'Must Fix',
                'Comment': (
                    f"Recovery Calc Est's {s['reimb_type']} total (${s['total_reimb']:,}) does not match "
                    f"Recovery Monthly's {s['reimb_type']} Total Reimbursement (${m_total:,}). "
                    "Likely a stale export - confirm both were generated from the same Kardin revision."
                ),
                'Status': 'Open', 'Source Check': 'Recovery Calc Est vs Monthly mismatch',
            })
    return findings


STOPWORDS = {'expense', 'expenses'}
GU_PCT_IN_GROUP_RE = re.compile(r'(\d+(?:\.\d+)?)\s*%\s*GU')


def _keywords(label):
    """Word-set for fuzzy matching, singular-stemmed and stripped of noise words.
    Fixed Factor Calcs uses short category names ("Water/Sewer"); the Gross Up
    Schedule uses full Kardin GL labels ("Utilities-Water/Sewer") - a plain
    substring or exact match fails on both the added prefix and pluralization
    ("Fee" vs "Fees"), so match on a stemmed word-subset instead."""
    words = re.findall(r'[a-z]+', label.lower())
    return {w[:-1] if w.endswith('s') and len(w) > 3 else w for w in words} - STOPWORDS


def _find_gu_match(ff_keywords, gu_blocks):
    best = None
    for b in gu_blocks:
        if not b['label']:
            continue
        gu_keywords = _keywords(b['label'])
        if ff_keywords and ff_keywords.issubset(gu_keywords):
            if best is None or len(gu_keywords) < len(_keywords(best['label'])):
                best = b  # prefer the tightest match if multiple GU lines qualify
    return best


def check_fixed_factor_vs_gross_up_schedule(fixed_factor_rows, gu_blocks, tolerance=5):
    findings = []
    gu_group_pcts = {}  # group name -> fraction, e.g. 0.95
    for b in gu_blocks:
        m = GU_PCT_IN_GROUP_RE.search(b['group'])
        if m:
            gu_group_pcts[b['group']] = float(m.group(1)) / 100.0
    for ff in fixed_factor_rows:
        if not ff['gl_description'] or ff['total_recoverable_expenses'] in (None, 0):
            continue
        candidate_blocks = [
            b for b in gu_blocks
            if ff['gu_group_pct'] is not None and gu_group_pcts.get(b['group']) == ff['gu_group_pct']
        ]
        gu = _find_gu_match(_keywords(str(ff['gl_description'])), candidate_blocks)
        if gu is None:
            continue
        diffs = []
        if gu['recoverable_expenses'] is not None:
            d = ff['total_recoverable_expenses'] - gu['recoverable_expenses']
            if abs(d) > tolerance:
                diffs.append(f"Total Recoverable Expenses: Fixed Factor Calcs ${ff['total_recoverable_expenses']:,.0f} "
                             f"vs Gross Up Schedule ${gu['recoverable_expenses']:,.0f} (diff ${d:+,.0f})")
        if gu['gu_amount'] is not None and ff['grossed_up_adjustment'] is not None:
            d2 = ff['grossed_up_adjustment'] - gu['gu_amount']
            if abs(d2) > tolerance:
                diffs.append(f"Gross-up Amount: Fixed Factor Calcs ${ff['grossed_up_adjustment']:,.0f} "
                             f"vs Gross Up Schedule ${gu['gu_amount']:,.0f} (diff ${d2:+,.0f})")
        if diffs:
            findings.append({
                'Report Section': '7. Recoveries & Fixed Factor',
                'GL Acct': gu.get('gl') or '', 'Line Item': ff['gl_description'], 'Budget Year': 'Next Year Budget',
                'Priority': 'Must Fix',
                'Comment': (
                    f"Fixed Factor Calcs workpaper doesn't tie to the Gross Up Schedule Kardin actually "
                    f"exported: " + '; '.join(diffs) + ". The workpaper is likely stale - re-check it "
                    "was updated after the last change to this GL line's budget."
                ),
                'Status': 'Open', 'Source Check': 'Fixed Factor Calcs stale vs Kardin',
            })
    return findings


def run(recovery_calc_est_pdf, recovery_monthly_pdf, gross_up_schedule_pdf, fixed_factor_xlsx):
    monthly_rows = parse_recovery_monthly(recovery_monthly_pdf)
    detail_rows, summary_rows = parse_recovery_calc_est(recovery_calc_est_pdf)
    gu_blocks, gu_group_totals, gu_page_info = parse_gross_up_schedule(gross_up_schedule_pdf)
    ff_rows = parse_fixed_factor_workbook(fixed_factor_xlsx)

    findings = []
    findings += check_gross_up_schedule_completeness(gu_page_info)
    findings += check_calc_est_vs_monthly_totals(summary_rows, monthly_rows)
    findings += check_fixed_factor_vs_gross_up_schedule(ff_rows, gu_blocks)

    stats = {
        'recovery_monthly_rows': len(monthly_rows),
        'recovery_calc_est_detail_rows': len(detail_rows),
        'recovery_calc_est_summary_rows': len(summary_rows),
        'gross_up_blocks': len(gu_blocks),
        'gross_up_page_info': gu_page_info,
        'fixed_factor_rows': len(ff_rows),
    }
    return {
        'monthly_rows': monthly_rows, 'detail_rows': detail_rows, 'summary_rows': summary_rows,
        'gu_blocks': gu_blocks, 'gu_group_totals': gu_group_totals, 'ff_rows': ff_rows,
        'findings': findings, 'stats': stats,
    }
