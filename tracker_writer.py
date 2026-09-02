"""
Insert parser findings into a copy of a GRP Budget Comment Tracker workbook,
preserving the sheet's section structure, styles, and COUNTIF formulas.

The tracker hardcodes absolute-row COUNTIF ranges per section (e.g.
"=COUNTIF(M22:M34,...)") and the Portfolio Summary hardcodes absolute
ranges per building tab. openpyxl's insert_rows() shifts cell content but
does NOT rewrite formula text or move merged-cell ranges, so both must be
recomputed/re-applied after insertion using a cumulative row-offset.
"""
import copy
import io
import re
from datetime import date

import openpyxl

HEADER_RE = re.compile(r'^\s*(\d+\.\s.+?)\s*$')
MARKER_RE = re.compile(r'Insert row above')

COLS = ['#', 'Report Section', 'GL Acct', 'Line Item', 'Budget Year', 'Month / Year',
        'Priority', 'Recurring', 'Comment By', 'Comment', 'Response By', 'PM Response',
        'Status', 'Notes']


class TrackerStructureError(Exception):
    pass


def find_sections(ws):
    """Return [{name, header_row, marker_row}] in top-to-bottom order."""
    headers = []
    markers = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if not isinstance(v, str):
            continue
        if MARKER_RE.search(v):
            markers[r] = True
        else:
            m = HEADER_RE.match(v)
            if m:
                headers.append((r, m.group(1)))
    sections = []
    marker_rows = sorted(markers.keys())
    for header_row, name in headers:
        later_markers = [m for m in marker_rows if m > header_row]
        if not later_markers:
            raise TrackerStructureError(f"No 'Insert row above' marker found after section '{name}' (row {header_row}).")
        sections.append({'name': name, 'header_row': header_row, 'marker_row': later_markers[0]})
    return sections


def copy_style(src_cell, dst_cell):
    dst_cell.font = copy.copy(src_cell.font)
    dst_cell.fill = copy.copy(src_cell.fill)
    dst_cell.border = copy.copy(src_cell.border)
    dst_cell.alignment = copy.copy(src_cell.alignment)
    dst_cell.number_format = src_cell.number_format


def next_number(ws, section_first_row, section_last_row):
    nums = []
    for r in range(section_first_row, section_last_row + 1):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, (int, float)):
            nums.append(int(v))
    return (max(nums) + 1) if nums else 1


def build_row_values(finding, seq_num, comment_by, today_str):
    return {
        1: seq_num,
        2: finding['Report Section'],
        3: finding['GL Acct'],
        4: finding['Line Item'],
        5: finding['Budget Year'],
        6: finding.get('Month / Year', 'Full Year'),
        7: finding['Priority'],
        8: 'No',  # 'Recurring' = seen in prior cycles; parser has no history to confirm
        9: comment_by,
        10: '[Automated QC] ' + finding['Comment'],
        11: None,
        12: None,
        13: finding['Status'],
        14: f'Auto-flagged by Kardin QC parser, {today_str}',
    }


def apply_findings_to_sheet(ws, findings, style_ref_row=8, comment_by='Ryan Walsh'):
    today_str = date.today().isoformat()
    sections = find_sections(ws)
    findings_by_section = {}
    for f in findings:
        findings_by_section.setdefault(f['Report Section'], []).append(f)

    unknown_sections = set(findings_by_section) - {s['name'] for s in sections}
    if unknown_sections:
        raise TrackerStructureError(
            f"Findings reference section(s) not found in this sheet: {sorted(unknown_sections)}. "
            f"Sheet sections are: {[s['name'] for s in sections]}"
        )

    insert_counts = {s['name']: len(findings_by_section.get(s['name'], [])) for s in sections}
    ordered_asc = sorted(sections, key=lambda s: s['header_row'])

    def offset_before_row(row):
        return sum(insert_counts[s['name']] for s in ordered_asc if s['header_row'] < row)

    # openpyxl's insert_rows() does NOT shift merged-cell ranges, which corrupts
    # the section-header/marker banner merges. Capture them, strip them, insert
    # rows, then re-merge at their offset-corrected positions.
    original_merges = [str(r) for r in ws.merged_cells.ranges]
    for r in original_merges:
        ws.unmerge_cells(r)

    ordered_desc = sorted(sections, key=lambda s: s['marker_row'], reverse=True)
    for s in ordered_desc:
        n = insert_counts[s['name']]
        if n:
            ws.insert_rows(s['marker_row'], amount=n)

    for r in original_merges:
        m = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', r)
        c1, row1, c2, row2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
        new_row1 = row1 + offset_before_row(row1)
        new_row2 = row2 + offset_before_row(row2)
        ws.merge_cells(f'{c1}{new_row1}:{c2}{new_row2}')

    offset_before = 0
    final_positions = {}
    for s in ordered_asc:
        n = insert_counts[s['name']]
        final_header = s['header_row'] + offset_before
        final_marker = s['marker_row'] + offset_before + n
        final_positions[s['name']] = {
            'header': final_header, 'marker': final_marker,
            'first_data_row': final_header + 1, 'insert_count': n,
        }
        offset_before += n

    style_cells = {c: ws.cell(row=style_ref_row, column=c) for c in range(1, 15)}
    for name, pos in final_positions.items():
        n = pos['insert_count']
        if not n:
            continue
        start_row = pos['marker'] - n
        seq = next_number(ws, pos['first_data_row'], start_row - 1)
        for i, finding in enumerate(findings_by_section[name]):
            r = start_row + i
            values = build_row_values(finding, seq + i, comment_by, today_str)
            for c in range(1, 15):
                cell = ws.cell(row=r, column=c)
                copy_style(style_cells[c], cell)
                cell.value = values[c]

    for name, pos in final_positions.items():
        header_cell = ws.cell(row=pos['header'], column=11)  # column K
        header_cell.value = f'=COUNTIF(M{pos["first_data_row"]}:M{pos["marker"]},"Open")&" open"'

    total_inserted = sum(insert_counts.values())
    last_section = ordered_asc[-1]['name']
    last_marker_final = final_positions[last_section]['marker']
    return total_inserted, last_marker_final


def merge_findings_into_tracker(tracker_file, sheet_name, findings, comment_by='Ryan Walsh'):
    """
    tracker_file: path or file-like object (e.g. Streamlit UploadedFile) of the
                  existing Comment Tracker .xlsx.
    sheet_name:   the building's tab name (must already exist in the workbook).
    findings:     list of finding dicts from kardin_parser checks.

    Returns (BytesIO of the updated .xlsx, total_inserted, warnings: list[str]).
    """
    wb = openpyxl.load_workbook(tracker_file, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise TrackerStructureError(
            f"Sheet '{sheet_name}' not found in tracker. Available sheets: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    orig_sections = find_sections(ws)
    orig_last_marker = max(s['marker_row'] for s in orig_sections)

    orig_buffer = 0
    ps_row = None
    if 'Portfolio Summary' in wb.sheetnames:
        ps = wb['Portfolio Summary']
        for r in range(1, ps.max_row + 1):
            if ps.cell(row=r, column=1).value == sheet_name:
                ps_row = r
                break
        if ps_row is not None:
            for col in range(2, ps.max_column + 1):
                formula = ps.cell(row=ps_row, column=col).value
                if isinstance(formula, str):
                    m = re.search(r"M\d+:M(\d+)", formula)
                    if m:
                        orig_buffer = int(m.group(1)) - orig_last_marker
                        break

    total_inserted, new_last_marker = apply_findings_to_sheet(ws, findings, comment_by=comment_by)

    warnings = []
    if ps_row is not None and total_inserted:
        ps = wb['Portfolio Summary']
        new_end = new_last_marker + orig_buffer
        for col in range(2, ps.max_column + 1):
            cell = ps.cell(row=ps_row, column=col)
            formula = cell.value
            if not isinstance(formula, str):
                continue
            cell.value = re.sub(r"M\d+:M\d+", lambda mm: re.sub(
                r'M(\d+):M(\d+)', lambda mm2: f"M{mm2.group(1)}:M{new_end}", mm.group(0)), formula)
    elif total_inserted and ps_row is None:
        warnings.append(
            "No matching row found in 'Portfolio Summary' for this building - "
            "portfolio-level totals will not include the new rows until you update it manually."
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, total_inserted, warnings
